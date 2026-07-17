import openpyxl
from xlsx_writer import escribir_xlsx


def test_escribe_hoja_lista_con_encabezados_y_datos(tmp_path):
    ruta = str(tmp_path / "salida.xlsx")
    lista = [
        {"nombre": "iPhone 13 128GB", "link": "http://x az", "pais": "🇺🇸", "precio": 665},
    ]
    reporte = {"filtrados": [], "duplicados_posibles": [], "dudas_precio": []}
    escribir_xlsx(lista, reporte, ruta)

    wb = openpyxl.load_workbook(ruta)
    assert "Lista" in wb.sheetnames
    assert "Reporte" in wb.sheetnames
    hoja = wb["Lista"]
    assert [c.value for c in hoja[1]] == ["Nombre", "Link Google Imágenes", "País", "Precio"]
    assert [c.value for c in hoja[2]] == ["iPhone 13 128GB", "http://x az", "🇺🇸", 665]


def test_reporte_incluye_secciones(tmp_path):
    ruta = str(tmp_path / "salida.xlsx")
    reporte = {
        "filtrados": [{"nombre": "iPhone (caja abollada)", "motivo": "caja abollada"}],
        "duplicados_posibles": [{"nombre_a": "iPhone 13 128GB", "nombre_b": "iPhone 13 256GB", "similitud": 0.75}],
        "dudas_precio": [{"texto": "Samsung consultar", "motivo": "sin costo"}],
    }
    escribir_xlsx([], reporte, ruta)
    wb = openpyxl.load_workbook(ruta)
    textos = [str(c.value) for fila in wb["Reporte"].iter_rows() for c in fila if c.value is not None]
    unido = " | ".join(textos)
    assert "Filtrados" in unido
    assert "caja abollada" in unido
    assert "Posibles duplicados" in unido
    assert "iPhone 13 256GB" in unido
    assert "Dudas de precio" in unido
    assert "Samsung consultar" in unido
